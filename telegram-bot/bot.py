"""
行政主管 Claude - Telegram Bot
透過 Telegram 呼叫 Claude AI 執行工作命令，並定時通知合約到期情況
"""

from __future__ import annotations

import os
import json
import logging
import datetime
from pathlib import Path
from collections import defaultdict
from dotenv import load_dotenv
import anthropic
from telegram import Update, ReplyKeyboardMarkup, KeyboardButton
from telegram.ext import (
    Application,
    CommandHandler,
    ConversationHandler,
    MessageHandler,
    filters,
    ContextTypes,
)

load_dotenv()

logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger(__name__)

TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN")
ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY")
ALLOWED_CHAT_IDS_RAW = os.getenv("ALLOWED_CHAT_IDS", "")
SYSTEM_PROMPT = os.getenv(
    "SYSTEM_PROMPT",
    "你是一位高效能的行政主管助理，負責協助處理各種工作任務。"
    "請用繁體中文回應，簡潔有力，直接提供可執行的建議與行動方案。"
    "你擅長：行程規劃、文件撰寫、資料整理、決策分析、任務追蹤。",
)

BASE_DIR = Path.home() / "Desktop" / "克勞德"

# Excel 來源設定（欄位名稱依老闆提供）
EXCEL_CONFIG = {
    "日日好日": {
        "file": BASE_DIR / "日日好日 總體客戶資料.xlsx",
        "col_name": os.getenv("RIHARI_COL_NAME", "名稱"),
        "col_id": os.getenv("RIHARI_COL_ID", "統編"),
        "col_period": os.getenv("RIHARI_COL_PERIOD", "合約結束日期"),
    },
    "日青": {
        "file": BASE_DIR / "日青營業登記.xlsx",
        "col_name": os.getenv("NICHINICHI_COL_NAME", "名稱"),
        "col_id": os.getenv("NICHINICHI_COL_ID", "統編"),
        "col_period": os.getenv("NICHINICHI_COL_PERIOD", "合約結束日期"),
    },
}

ALLOWED_CHAT_IDS: set[int] = set()
if ALLOWED_CHAT_IDS_RAW.strip():
    for _cid in ALLOWED_CHAT_IDS_RAW.split(","):
        _cid = _cid.strip()
        if _cid:
            ALLOWED_CHAT_IDS.add(int(_cid))

conversation_history: dict[int, list[dict]] = defaultdict(list)
client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

ASK_CATEGORY, ASK_NAME, ASK_DATE, ASK_NOTES = range(4)

MAIN_KEYBOARD = ReplyKeyboardMarkup(
    [
        [KeyboardButton("🏢 日日好日合約"), KeyboardButton("🌿 日青合約")],
        [KeyboardButton("➕ 新增合約"), KeyboardButton("🗑️ 刪除合約")],
        [KeyboardButton("⚠️ 檢查到期"), KeyboardButton("🗂️ 清除對話")],
    ],
    resize_keyboard=True,
)

CATEGORY_KEYBOARD = ReplyKeyboardMarkup(
    [
        [KeyboardButton("🏢 日日好日"), KeyboardButton("🌿 日青")],
        [KeyboardButton("❌ 取消")],
    ],
    resize_keyboard=True,
)


# ── 民國日期解析 ──────────────────────────────────────────────

def parse_roc_date(s: str) -> datetime.date | None:
    """解析民國日期字串 YYYMMDD → datetime.date"""
    s = str(s).strip().replace("/", "").replace(".", "")
    # 可能是 7 碼（民國）
    if len(s) == 7:
        try:
            year = int(s[:3]) + 1911
            month = int(s[3:5])
            day = int(s[5:7])
            return datetime.date(year, month, day)
        except (ValueError, TypeError):
            return None
    return None


def parse_roc_period(value) -> tuple[datetime.date | None, datetime.date | None]:
    """解析 'YYYMMDD-YYYMMDD' 格式，回傳 (起始日, 結束日)"""
    s = str(value).strip()
    if "-" in s:
        parts = s.split("-", 1)
        if len(parts) == 2:
            return parse_roc_date(parts[0]), parse_roc_date(parts[1])
    # 只有單一日期視為結束日
    return None, parse_roc_date(s)


def to_roc_str(d: datetime.date | None) -> str:
    """datetime.date → 民國年顯示字串"""
    if not d:
        return "—"
    return f"民國{d.year - 1911}年{d.month}月{d.day}日"


# ── Excel 讀取 ────────────────────────────────────────────────

def read_excel_contracts(category: str) -> tuple[list[dict], str]:
    """
    讀取指定業務的 Excel 合約資料。
    回傳 (合約清單, 錯誤訊息)。
    """
    cfg = EXCEL_CONFIG.get(category)
    if not cfg:
        return [], f"找不到「{category}」的設定"

    filepath: Path = cfg["file"]
    if not filepath.exists():
        return [], f"找不到 Excel 檔案：\n{filepath}\n\n請確認桌面「克勞德」資料夾內有此檔案。"

    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active

        # 自動掃描前 5 行，找到含有欄位名稱的那一行
        header_row_idx = None
        headers = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
            row_strs = [str(v).strip() if v is not None else "" for v in row]
            if cfg["col_name"] in row_strs or cfg["col_period"] in row_strs:
                header_row_idx = row_idx
                headers = row_strs
                break

        if not headers:
            # 回報前 5 行內容供除錯
            sample = []
            for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
                sample.append(", ".join(str(v) for v in row if v is not None))
            return [], (
                f"【{category}】找不到欄位標題行。\n"
                f"Excel 前三行內容：\n" + "\n".join(sample)
            )

        data_start_row = header_row_idx + 1

        def find_col(col_name: str) -> int | None:
            try:
                return headers.index(col_name)
            except ValueError:
                return None

        name_idx = find_col(cfg["col_name"])
        id_idx = find_col(cfg["col_id"])
        period_idx = find_col(cfg["col_period"])

        if name_idx is None:
            return [], (
                f"【{category}】找不到「{cfg['col_name']}」欄位。\n"
                f"Excel 現有欄位：{', '.join(h for h in headers if h)}"
            )
        if period_idx is None:
            return [], (
                f"【{category}】找不到「{cfg['col_period']}」欄位。\n"
                f"Excel 現有欄位：{', '.join(h for h in headers if h)}"
            )

        contracts = []
        for row in ws.iter_rows(min_row=data_start_row, values_only=True):
            if all(v is None for v in row):
                continue
            name = str(row[name_idx]).strip() if row[name_idx] is not None else ""
            if not name or name.lower() == "none":
                continue

            period_val = row[period_idx] if period_idx < len(row) else None
            if not period_val:
                continue

            start_date, end_date = parse_roc_period(period_val)

            unified_id = ""
            if id_idx is not None and id_idx < len(row) and row[id_idx] is not None:
                unified_id = str(row[id_idx]).strip()

            contracts.append({
                "name": name,
                "unified_id": unified_id,
                "contract_start_date": start_date.isoformat() if start_date else "",
                "contract_end_date": end_date.isoformat() if end_date else "",
                "_category": category,
                "_start_date_obj": start_date,
                "_end_date_obj": end_date,
            })

        return contracts, ""

    except Exception as e:
        logger.exception("讀取 Excel 失敗 [%s]", category)
        return [], f"讀取 Excel 時發生錯誤：{e}"


def get_expiring_from_excel(category: str, within_days: int) -> tuple[list[dict], str]:
    """取得指定天數內到期的合約，回傳 (到期清單, 錯誤訊息)"""
    contracts, err = read_excel_contracts(category)
    if err:
        return [], err
    today = datetime.date.today()
    deadline = today + datetime.timedelta(days=within_days)
    result = []
    for c in contracts:
        end = c.get("_end_date_obj")
        if end and today <= end <= deadline:
            result.append({**c, "days_left": (end - today).days})
    result.sort(key=lambda x: x["days_left"])
    return result, ""


# ── 工具函式 ─────────────────────────────────────────────────

def is_authorized(chat_id: int) -> bool:
    if not ALLOWED_CHAT_IDS:
        return True
    return chat_id in ALLOWED_CHAT_IDS


def days_status(days_left: int) -> str:
    if days_left < 0:
        return f"已過期 {abs(days_left)} 天"
    if days_left == 0:
        return "今天到期！"
    if days_left <= 7:
        return f"還有 {days_left} 天 🔴"
    if days_left <= 21:
        return f"還有 {days_left} 天 🟡"
    return f"還有 {days_left} 天"


async def send_long(update_or_bot, chat_id: int, text: str, **kwargs) -> None:
    if hasattr(update_or_bot, "message"):
        for i in range(0, len(text), 4096):
            await update_or_bot.message.reply_text(text[i: i + 4096], **kwargs)
    else:
        for i in range(0, len(text), 4096):
            await update_or_bot.send_message(chat_id=chat_id, text=text[i: i + 4096], **kwargs)


def format_contract_entry(c: dict, idx: int) -> str:
    start = to_roc_str(c.get("_start_date_obj"))
    end = to_roc_str(c.get("_end_date_obj"))
    days_left = c.get("days_left", "")
    status = f"（還有 {days_left} 天）" if days_left != "" else ""
    return (
        f"{idx}. {c.get('name', '—')}\n"
        f"   統編：{c.get('unified_id') or '—'}\n"
        f"   合約期間：{start} ~ {end}{status}"
    )


def format_contracts_list(category: str) -> str:
    contracts, err = read_excel_contracts(category)
    if err:
        return f"【{category}】⚠️ {err}"
    if not contracts:
        return f"【{category}】Excel 中目前沒有資料。"

    today = datetime.date.today()
    # 依到期日排序
    contracts.sort(key=lambda c: c.get("contract_end_date") or "")

    lines = [f"【{category}】合約清單（共 {len(contracts)} 筆）\n"]
    for i, c in enumerate(contracts, 1):
        end = c.get("_end_date_obj")
        if end:
            days_left = (end - today).days
            status = days_status(days_left)
        else:
            status = "—"
        start = to_roc_str(c.get("_start_date_obj"))
        end_str = to_roc_str(end)
        lines.append(
            f"{i}. {c.get('name', '—')}\n"
            f"   統編：{c.get('unified_id') or '—'}\n"
            f"   期間：{start} ~ {end_str}（{status}）"
        )
    return "\n".join(lines)


# ── 一般指令 ──────────────────────────────────────────────────

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not is_authorized(update.effective_chat.id):
        await update.message.reply_text("抱歉，您沒有使用此機器人的權限。")
        return
    await update.message.reply_text(
        "您好！我是您的行政主管助理 Claude。\n\n"
        "直接輸入工作指令，或點選下方按鈕操作合約管理。",
        reply_markup=MAIN_KEYBOARD,
    )


async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not is_authorized(update.effective_chat.id):
        await update.message.reply_text("抱歉，您沒有使用此機器人的權限。")
        return
    await update.message.reply_text(
        "按鈕功能說明：\n\n"
        "🏢 日日好日合約 — 查看日日好日完整合約清單\n"
        "🌿 日青合約 — 查看日青完整合約清單\n"
        "⚠️ 檢查到期 — 本週 + 三週內即將到期清單\n"
        "➕ 新增合約 — 手動新增單筆合約\n"
        "🗑️ 刪除合約 — 刪除手動新增的合約\n"
        "🗂️ 清除對話 — 清除 AI 對話記錄\n\n"
        "每週一 10:00 自動推送到期通知。\n"
        f"Excel 資料來源：\n"
        f"  日日好日：日日好日 總體客戶資料.xlsx\n"
        f"  日青：日青營業登記.xlsx",
        reply_markup=MAIN_KEYBOARD,
    )


async def id_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user = update.effective_user
    await update.message.reply_text(
        f"您的 Chat ID：{update.effective_chat.id}\n使用者名稱：{user.full_name}"
    )


async def clear_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not is_authorized(update.effective_chat.id):
        await update.message.reply_text("抱歉，您沒有使用此機器人的權限。")
        return
    conversation_history[update.effective_chat.id].clear()
    await update.message.reply_text("對話記錄已清除。", reply_markup=MAIN_KEYBOARD)


# ── 查看合約 ──────────────────────────────────────────────────

async def list_rihari(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not is_authorized(update.effective_chat.id):
        await update.message.reply_text("抱歉，您沒有使用此機器人的權限。")
        return
    await send_long(update, update.effective_chat.id, format_contracts_list("日日好日"), reply_markup=MAIN_KEYBOARD)


async def list_nichinichi(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not is_authorized(update.effective_chat.id):
        await update.message.reply_text("抱歉，您沒有使用此機器人的權限。")
        return
    await send_long(update, update.effective_chat.id, format_contracts_list("日青"), reply_markup=MAIN_KEYBOARD)


# ── 檢查到期 ──────────────────────────────────────────────────

async def check_contracts(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not is_authorized(update.effective_chat.id):
        await update.message.reply_text("抱歉，您沒有使用此機器人的權限。")
        return
    await _send_contract_report(context.bot, update.effective_chat.id)


# ── 刪除合約（手動新增的才能刪）────────────────────────────────

def load_manual_contracts(category: str) -> list[dict]:
    f = BASE_DIR / f"{category}續約合約" / "contracts.json"
    if not f.exists():
        return []
    with open(f, encoding="utf-8") as fp:
        return json.load(fp)


def save_manual_contracts(category: str, contracts: list[dict]) -> None:
    folder = BASE_DIR / f"{category}續約合約"
    folder.mkdir(parents=True, exist_ok=True)
    with open(folder / "contracts.json", "w", encoding="utf-8") as fp:
        json.dump(contracts, fp, ensure_ascii=False, indent=2)


async def remove_contract(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not is_authorized(update.effective_chat.id):
        await update.message.reply_text("抱歉，您沒有使用此機器人的權限。")
        return

    all_contracts = []
    for cat in ("日日好日", "日青"):
        for c in load_manual_contracts(cat):
            all_contracts.append({**c, "_category": cat})

    if not all_contracts:
        await update.message.reply_text("目前沒有手動新增的合約記錄。", reply_markup=MAIN_KEYBOARD)
        return

    lines = ["請回傳要刪除的合約編號：\n"]
    for i, c in enumerate(all_contracts, 1):
        lines.append(f"{i}. [{c['_category']}] {c.get('name', '—')}｜{c.get('contract_end_date', '—')}")
    lines.append("\n直接回傳數字即可，例如：2")
    await update.message.reply_text("\n".join(lines))
    context.user_data["awaiting_remove"] = True
    context.user_data["remove_pool"] = all_contracts


# ── 新增合約（ConversationHandler）────────────────────────────

async def addcontract_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if not is_authorized(update.effective_chat.id):
        await update.message.reply_text("抱歉，您沒有使用此機器人的權限。")
        return ConversationHandler.END
    context.user_data["new_contract"] = {}
    await update.message.reply_text(
        "新增合約\n\n請選擇業務：",
        reply_markup=CATEGORY_KEYBOARD,
    )
    return ASK_CATEGORY


async def got_category(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text == "❌ 取消":
        await update.message.reply_text("已取消。", reply_markup=MAIN_KEYBOARD)
        return ConversationHandler.END
    if "日日好日" in text:
        context.user_data["new_contract"]["_category"] = "日日好日"
    elif "日青" in text:
        context.user_data["new_contract"]["_category"] = "日青"
    else:
        await update.message.reply_text("請選擇「🏢 日日好日」或「🌿 日青」：", reply_markup=CATEGORY_KEYBOARD)
        return ASK_CATEGORY
    await update.message.reply_text("請輸入合約對象的姓名：")
    return ASK_NAME


async def got_name(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data["new_contract"]["name"] = update.message.text.strip()
    await update.message.reply_text("請輸入合約到期日（格式：YYYY-MM-DD，例如 2026-12-31）：")
    return ASK_DATE


async def got_date(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    raw = update.message.text.strip()
    try:
        datetime.date.fromisoformat(raw)
    except ValueError:
        await update.message.reply_text("日期格式不正確，請重新輸入（格式：YYYY-MM-DD）：")
        return ASK_DATE
    context.user_data["new_contract"]["contract_end_date"] = raw
    await update.message.reply_text("請輸入備註（直接輸入「無」跳過）：")
    return ASK_NOTES


async def got_notes(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    notes = update.message.text.strip()
    data = context.user_data.pop("new_contract")
    category = data.pop("_category")
    data["notes"] = "" if notes == "無" else notes

    contracts = load_manual_contracts(category)
    contracts.append(data)
    save_manual_contracts(category, contracts)

    await update.message.reply_text(
        f"合約已手動新增到【{category}】！\n\n"
        f"姓名：{data['name']}\n"
        f"到期日：{data['contract_end_date']}\n"
        f"備註：{data.get('notes') or '無'}",
        reply_markup=MAIN_KEYBOARD,
    )
    return ConversationHandler.END


async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data.pop("new_contract", None)
    await update.message.reply_text("已取消。", reply_markup=MAIN_KEYBOARD)
    return ConversationHandler.END


# ── 定時通知邏輯 ──────────────────────────────────────────────

async def _send_contract_report(bot, chat_id: int) -> None:
    today = datetime.date.today()
    lines = [f"合約到期通知｜{today.strftime('%Y/%m/%d')}\n"]
    found_any = False

    category_labels = {
        "日日好日": "【日日好日】小鐘提回報",
        "日青": "【日青】小青提回報",
    }

    for category, label in category_labels.items():
        # 本週到期（7天內）
        week_list, err_w = get_expiring_from_excel(category, 7)
        # 三週內到期（8~21天）
        three_week_list, err_3 = get_expiring_from_excel(category, 21)
        three_week_list = [c for c in three_week_list if c["days_left"] > 7]

        if err_w and err_3:
            lines.append(f"{label}：⚠️ {err_w}\n")
            continue

        if not week_list and not three_week_list:
            continue

        found_any = True
        lines.append(f"━━━ {label} ━━━")

        if week_list:
            lines.append("🔴 本週到期：")
            for idx, c in enumerate(week_list, 1):
                lines.append(format_contract_entry(c, idx))
        else:
            lines.append("🔴 本週到期：無")

        if three_week_list:
            lines.append("\n🟡 三週內即將到期：")
            for idx, c in enumerate(three_week_list, 1):
                lines.append(format_contract_entry(c, idx))
        else:
            lines.append("🟡 三週內即將到期：無")

        lines.append("")

    if not found_any:
        await bot.send_message(
            chat_id=chat_id,
            text=f"合約到期通知｜{today.strftime('%Y/%m/%d')}\n\n本週及三週內沒有即將到期的合約。",
        )
        return

    await send_long(bot, chat_id, "\n".join(lines))


async def weekly_contract_job(context: ContextTypes.DEFAULT_TYPE) -> None:
    for chat_id in ALLOWED_CHAT_IDS:
        try:
            await _send_contract_report(context.bot, chat_id)
        except Exception as e:
            logger.error("Failed to send weekly report to %s: %s", chat_id, e)


# ── 一般訊息處理 ──────────────────────────────────────────────

BUTTON_MAP = {
    "🏢 日日好日合約": list_rihari,
    "🌿 日青合約": list_nichinichi,
    "⚠️ 檢查到期": check_contracts,
    "🗑️ 刪除合約": remove_contract,
    "🗂️ 清除對話": clear_command,
}


async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    chat_id = update.effective_chat.id
    if not is_authorized(chat_id):
        await update.message.reply_text("抱歉，您沒有使用此機器人的權限。")
        return

    text = update.message.text or ""

    if text in BUTTON_MAP:
        await BUTTON_MAP[text](update, context)
        return

    if context.user_data.get("awaiting_remove"):
        pool = context.user_data.get("remove_pool", [])
        try:
            idx = int(text.strip()) - 1
            if 0 <= idx < len(pool):
                target = pool[idx]
                cat = target["_category"]
                contracts = load_manual_contracts(cat)
                contracts = [
                    c for c in contracts
                    if not (c.get("name") == target.get("name") and
                            c.get("contract_end_date") == target.get("contract_end_date"))
                ]
                save_manual_contracts(cat, contracts)
                context.user_data.pop("awaiting_remove", None)
                context.user_data.pop("remove_pool", None)
                await update.message.reply_text(
                    f"已刪除：[{cat}] {target.get('name', '—')}",
                    reply_markup=MAIN_KEYBOARD,
                )
            else:
                await update.message.reply_text(f"編號無效，請輸入 1 到 {len(pool)} 之間的數字。")
        except ValueError:
            await update.message.reply_text("請輸入數字編號。")
        return

    if not text:
        return

    await context.bot.send_chat_action(chat_id=chat_id, action="typing")
    conversation_history[chat_id].append({"role": "user", "content": text})
    history = conversation_history[chat_id][-40:]

    try:
        response = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=2048,
            system=SYSTEM_PROMPT,
            messages=history,
        )
        assistant_reply = response.content[0].text
        conversation_history[chat_id].append({"role": "assistant", "content": assistant_reply})
        await send_long(update, chat_id, assistant_reply, reply_markup=MAIN_KEYBOARD)
    except anthropic.APIError as e:
        logger.error("Anthropic API error: %s", e)
        await update.message.reply_text(f"Claude API 發生錯誤：{e}")
    except Exception as e:
        logger.error("Unexpected error: %s", e)
        await update.message.reply_text("發生未預期的錯誤，請稍後再試。")


# ── 主程式 ────────────────────────────────────────────────────

def main() -> None:
    if not TELEGRAM_BOT_TOKEN:
        raise ValueError("請在 .env 設定 TELEGRAM_BOT_TOKEN")
    if not ANTHROPIC_API_KEY:
        raise ValueError("請在 .env 設定 ANTHROPIC_API_KEY")

    app = Application.builder().token(TELEGRAM_BOT_TOKEN).build()

    add_contract_handler = ConversationHandler(
        entry_points=[
            CommandHandler("addcontract", addcontract_start),
            MessageHandler(filters.Regex("^➕ 新增合約$"), addcontract_start),
        ],
        states={
            ASK_CATEGORY: [MessageHandler(filters.TEXT & ~filters.COMMAND, got_category)],
            ASK_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, got_name)],
            ASK_DATE: [MessageHandler(filters.TEXT & ~filters.COMMAND, got_date)],
            ASK_NOTES: [MessageHandler(filters.TEXT & ~filters.COMMAND, got_notes)],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
    )

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("help", help_command))
    app.add_handler(CommandHandler("clear", clear_command))
    app.add_handler(CommandHandler("id", id_command))
    app.add_handler(CommandHandler("checkcontracts", check_contracts))
    app.add_handler(add_contract_handler)
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))

    if ALLOWED_CHAT_IDS and app.job_queue:
        app.job_queue.run_daily(
            weekly_contract_job,
            time=datetime.time(hour=2, minute=0, tzinfo=datetime.timezone.utc),
            days=(0,),
            name="weekly_contract_check",
        )
        logger.info("每週一 10:00（台灣時間）合約到期通知已設定")

    logger.info("行政主管 Claude Bot 啟動中...")
    app.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()
